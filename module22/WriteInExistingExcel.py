import openpyxl

# Load the Excel sheet
workbook = openpyxl.load_workbook("C:\\Users\\Piyush Verma\\OneDrive\\Documents\\readWriteExcelInPython\\writeExcel.xlsx")

# sheet
sheet1 = workbook['details']

# total numbers of rows
rowCount = sheet1.max_row

# total numbers of columns
columnCount = sheet1.max_column

# data
sheet1.cell(row=rowCount+1, column=1, value='Abhay')
sheet1.cell(row=rowCount+1, column=2, value='Kumar')
sheet1.cell(row=rowCount+1, column=3, value='Gupta')
sheet1.cell(row=rowCount+1, column=4, value=2222222)

# perform the write action
workbook.save("C:\\Users\\Piyush Verma\\OneDrive\\Documents\\readWriteExcelInPython\\writeExcel.xlsx")

# print the message
print("Data has Successfully written...")

# close the workbook
workbook.close()