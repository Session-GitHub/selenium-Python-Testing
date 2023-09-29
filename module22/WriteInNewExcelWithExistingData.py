import openpyxl

# Load the Excel sheet
workbook = openpyxl.load_workbook("C:\\Users\\Piyush Verma\\OneDrive\\Documents\\readWriteExcelInPython\\readExcel.xlsx")

# sheet
sheet1 = workbook['details']

# total numbers of rows
rowCount = sheet1.max_row
# total numbers of columns
columnCount = sheet1.max_column

# data
sheet1.cell(row=rowCount+1, column=1, value='Manoj')
sheet1.cell(row=rowCount+1, column=2, value='Pal')
sheet1.cell(row=rowCount+1, column=3, value='Singh')
sheet1.cell(row=rowCount+1, column=4, value=989999999)

workbook.save("C:\\Users\\Piyush Verma\\OneDrive\\Documents\\readWriteExcelInPython\\writeNewExcel.xlsx")

# print a message
print("Data has successfully written...")

# close the workbook
workbook.close()