import openpyxl

# Load the Excel sheet
workbook = openpyxl.load_workbook("C:\\Users\\Piyush Verma\\OneDrive\\Documents\\readWriteExcelInPython\\readExcel.xlsx")

# sheet
sheet1 = workbook['marks']

# print the total numbers of rows
rowCount = sheet1.max_row

# print the total numbers of columns
columnCount = sheet1.max_column

# print the data
for i in range(1, rowCount+1):
    for j in range(1, columnCount+1):
        data = sheet1.cell(i, j).value
        print(data)

# close the workbook
workbook.close()