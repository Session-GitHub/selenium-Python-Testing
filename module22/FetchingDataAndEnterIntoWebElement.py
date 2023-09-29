import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

# Creating the object of service class
service_obj = Service()

# initializing the Chrome browser
driver = webdriver.Chrome(service=service_obj)

# maximizing browser, apply implicitly wait & PageLoadTimeOut
driver.maximize_window()
driver.implicitly_wait(30)
driver.set_page_load_timeout(50)

# Apply the static wait
time.sleep(5)

# navigating to the URL
driver.get("https://demo.midtrans.com/")

# Apply the static wait
time.sleep(5)

# click on buyNow Button
driver.find_element(By.XPATH, "//a[@class='btn buy']").click()

# Apply the static wait
time.sleep(5)

nameElement = driver.find_element(By.XPATH, "(//table/tbody//tr//td/input)[2]")
emailElement = driver.find_element(By.XPATH, "(//table/tbody//tr//td/input)[3]")
phoneNoElement = driver.find_element(By.XPATH, "(//table/tbody//tr//td/input)[4]")
cityElement = driver.find_element(By.XPATH, "(//table/tbody//tr//td/input)[5]")
addressElement = driver.find_element(By.XPATH, "//table/tbody//tr//td/textarea")
postalCodeElement = driver.find_element(By.XPATH, "(//table/tbody//tr//td/input)[6]")

# Load the Excel sheet
workbook = openpyxl.load_workbook("C:\\Users\\Piyush Verma\\OneDrive\\Documents\\readWriteExcelInPython\\midtransData.xlsx")

# sheet
sheet1 = workbook['midtrans']

# taking data from Excel file
nameValue = sheet1.cell(1, 2).value
emailValue = sheet1.cell(2, 2).value
phoneNoValue = sheet1.cell(3, 2).value
cityValue = sheet1.cell(4, 2).value
addressValue = sheet1.cell(5, 2).value
postalCodeValue = sheet1.cell(6, 2).value

# enter the name
nameElement.clear()
nameElement.send_keys(nameValue)
time.sleep(3)

# enter the email
emailElement.clear()
emailElement.send_keys(emailValue)
time.sleep(3)

# enter the phone number
phoneNoElement.clear()
phoneNoElement.send_keys(phoneNoValue)
time.sleep(3)

# enter the city
cityElement.clear()
cityElement.send_keys(cityValue)
time.sleep(3)

# enter the address
addressElement.clear()
addressElement.send_keys(addressValue)
time.sleep(3)

# enter the postal code
postalCodeElement.clear()
postalCodeElement.send_keys(postalCodeValue)

# Apply the static wait
time.sleep(5)

# close the browser
driver.close()