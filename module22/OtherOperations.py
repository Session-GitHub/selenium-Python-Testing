import openpyxl

# Load the Excel sheet
workbook = openpyxl.load_workbook("C:\\Users\\Piyush Verma\\OneDrive\\Documents\\readWriteExcelInPython\\readExcel.xlsx")

# print the types of workbook
# print(type(workbook))

# all sheets name in "readExcel.xlsx"
# sheetNames = workbook.sheetnames
# print(sheetNames)

# print the name of active sheet
activeSheet = workbook.active.title
print(activeSheet)

# read particular sheet
sheet1 = workbook['details']
sheet2 = workbook['marks']
sheet3 = workbook['midtrans']

# option-1
# read cell data of details sheet
data1 = sheet1['B4'].value
print(data1)

# option-2
# print the cell data using the sheet name
data2 = workbook['details']['A6'].value
print(data2)

# option-3
# print the cell value using cell(row,column)
data3 = sheet1.cell(row=8, column=1).value
print(data3)

# option-4
# print the cell value using cell index
data4 = sheet1.cell(6, 1).value
print(data4)

# print the total numbers of rows
rowCount = sheet1.max_row
print("Total number of rows in 'details' sheet : ", rowCount)

# print the total numbers of columns
columnCount = sheet1.max_column
print("Total number of columns in 'details' sheet : ", columnCount)

# close the workbook
workbook.close()