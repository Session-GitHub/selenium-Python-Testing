import openpyxl

# Load the Excel sheet
workbook = openpyxl.load_workbook("C:\\Users\\Piyush Verma\\OneDrive\\Documents\\readWriteExcelInPython\\readExcel.xlsx")

# sheet
sheet1 = workbook['details']

# declare an empty dictionary
Dict = {}

# take the firstName
firstName = "Yatendra"

# total numbers of rows
rowCount = sheet1.max_row
# total numbers of columns
columnCount = sheet1.max_column

# Dictionary example
# dict_data = [{"FirstName" : "Piyush"}, {"LastName" : "Verma"}]

# print the details of "Piyush" in the form of Dictionary
for i in range(1, rowCount+1):
    if sheet1.cell(row=i, column=1).value == firstName:

        for j in range(2, columnCount+1):
            Dict[sheet1.cell(row=1, column=j).value] = sheet1.cell(i, j).value

# prin the data of "Piyush" in the form of Dictionary
print(Dict)

# print a message
print("Successfully Fetched...")

# close workbook
workbook.close()